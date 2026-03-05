import sys
from pathlib import Path

# Try to read the Excel file using different methods
excel_path = Path(r"C:\Users\krush\OneDrive - Walmart Inc\Documents\VSCode\Activity_Hub\Store Support\Projects\JobCodes-teaming\Job Codes\Job_Code_Master_Table.xlsx")

print(f"Excel file exists: {excel_path.exists()}")
print(f"File size: {excel_path.stat().st_size} bytes")

# Try using openpyxl directly
try:
    from openpyxl import load_workbook
    wb = load_workbook(str(excel_path))
    print(f"\nSheet names: {wb.sheetnames}")
    
    # Get first sheet
    ws = wb.active
    print(f"Active sheet: {ws.title}")
    print(f"Dimensions: {ws.dimensions}")
    
    # Read first few rows to see structure
    print("\n" + "="*80)
    print("FIRST ROW (HEADERS):")
    print("="*80)
    headers = []
    for col_idx, cell in enumerate(ws[1], 1):
        headers.append(cell.value)
        if col_idx <= 30:
            print(f"Col {col_idx}: {cell.value}")
    
    print(f"\nTotal columns: {len(headers)}")
    
    # Look for job code and category columns
    print("\n" + "="*80)
    print("RELEVANT COLUMNS (Job/Code/Category/Role):")
    print("="*80)
    relevant_cols = {}
    for idx, header in enumerate(headers):
        if header and any(keyword in str(header).lower() for keyword in ['job', 'code', 'category', 'role', 'salary', 'hourly', 'level', 'family']):
            relevant_cols[idx+1] = header
            print(f"Col {idx+1}: {header}")
    
    # Read sample data rows
    print("\n" + "="*80)
    print("SAMPLE DATA (First 10 rows):")
    print("="*80)
    for row_idx in range(2, 12):
        print(f"\nRow {row_idx}:")
        for col_idx in sorted(relevant_cols.keys())[:10]:  # Show first 10 relevant columns
            cell = ws.cell(row=row_idx, column=col_idx)
            print(f"  {relevant_cols[col_idx]}: {cell.value}")
    
    # Count rows
    max_row = ws.max_row
    print(f"\n" + "="*80)
    print(f"Total data rows: {max_row - 1}")  # Subtract header
    print("="*80)
    
except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
