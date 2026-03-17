"""
Query TMS Data (3).xlsx to see actual columns
"""
import openpyxl
from openpyxl import load_workbook

tms_file = r"c:\Users\krush\OneDrive - Walmart Inc\Documents\VSCode\Activity_Hub\Store Support\Projects\JobCodes-teaming\Teaming\TMS Data (3).xlsx"

print("="*80)
print("TMS DATA (3).XLSX - COLUMN ANALYSIS")
print("="*80)

try:
    # Load the workbook
    wb = load_workbook(tms_file)
    ws = wb.active
    
    # Get dimensions
    max_row = ws.max_row
    max_col = ws.max_column
    
    print(f"\n✓ File loaded: {tms_file}")
    print(f"\nDimensions: {max_row} rows × {max_col} columns")
    
    print(f"\n📋 COLUMN HEADERS (Row 1):")
    print("-" * 80)
    
    columns = []
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        col_name = cell.value
        columns.append(col_name)
        print(f"  Column {col_idx:2d}: {col_name}")
    
    print(f"\n📊 FIRST 5 DATA ROWS:")
    print("-" * 80)
    
    for row_idx in range(2, min(7, max_row + 1)):
        row_data = []
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            row_data.append(str(cell.value)[:15])  # Truncate for readability
        print(f"  Row {row_idx}: {' | '.join(row_data)}")
    
    print(f"\n✓ Data Types:")
    print("-" * 80)
    for col_idx, col_name in enumerate(columns, 1):
        sample_cell = ws.cell(row=2, column=col_idx)
        print(f"  {col_name}: {sample_cell.data_type} (sample: {sample_cell.value})")
    
    print(f"\n" + "="*80)
    print(f"SUMMARY: {max_col} columns total")
    print("="*80)
    
    wb.close()
    
except Exception as e:
    print(f"\n✗ Error: {e}")
    import traceback
    traceback.print_exc()
