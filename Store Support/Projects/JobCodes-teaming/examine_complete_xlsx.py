"""
Examine Job_Code_Master_Complete.xlsx vs Job_Code_Master_Table.xlsx
Compare structure, find additional mapping columns, and identify potential improvements
"""

import openpyxl
from pathlib import Path
from collections import defaultdict

# Find the Excel files
base_path = Path(r"c:\Users\krush\OneDrive - Walmart Inc\Documents\VSCode\Activity_Hub\Store Support\Projects\JobCodes-teaming\Job Codes")
complete_file = base_path / "Job_Code_Master_Complete.xlsx"
table_file = base_path / "Job_Code_Master_Table.xlsx"

print("=" * 80)
print("EXCEL FILE STRUCTURE COMPARISON")
print("=" * 80)

# Check if files exist
if not complete_file.exists():
    print(f"❌ Not found: {complete_file}")
else:
    print(f"✅ Found: {complete_file.name}")

if not table_file.exists():
    print(f"❌ Not found: {table_file}")
else:
    print(f"✅ Found: {table_file.name}")

print("\n" + "=" * 80)
print("ANALYZING JOB_CODE_MASTER_COMPLETE.xlsx")
print("=" * 80)

try:
    wb_complete = openpyxl.load_workbook(complete_file)
    print(f"\n📊 Sheet Names: {wb_complete.sheetnames}")
    
    ws_complete = wb_complete.active
    print(f"📄 Active Sheet: {ws_complete.title}")
    print(f"📈 Dimensions: {ws_complete.dimensions}")
    
    # Get headers
    headers_complete = []
    for cell in ws_complete[1]:
        if cell.value:
            headers_complete.append(cell.value)
    
    print(f"\n📋 Columns ({len(headers_complete)}):")
    for i, header in enumerate(headers_complete, 1):
        print(f"   [{i}] {header}")
    
    # Count rows with data
    row_count_complete = ws_complete.max_row - 1  # Exclude header
    print(f"\n📊 Total Data Rows: {row_count_complete}")
    
    # Show first 5 rows as examples
    print(f"\n📝 First 5 Data Rows:")
    for row_idx in range(2, min(7, ws_complete.max_row + 1)):
        row_data = []
        for col_idx in range(1, len(headers_complete) + 1):
            cell_value = ws_complete.cell(row_idx, col_idx).value
            row_data.append(str(cell_value)[:30] if cell_value else "")
        print(f"   Row {row_idx}: {row_data}")

except Exception as e:
    print(f"❌ Error reading Complete file: {e}")

print("\n" + "=" * 80)
print("ANALYZING JOB_CODE_MASTER_TABLE.xlsx")
print("=" * 80)

try:
    wb_table = openpyxl.load_workbook(table_file)
    print(f"\n📊 Sheet Names: {wb_table.sheetnames}")
    
    ws_table = wb_table.active
    print(f"📄 Active Sheet: {ws_table.title}")
    print(f"📈 Dimensions: {ws_table.dimensions}")
    
    # Get headers
    headers_table = []
    for cell in ws_table[1]:
        if cell.value:
            headers_table.append(cell.value)
    
    print(f"\n📋 Columns ({len(headers_table)}):")
    for i, header in enumerate(headers_table, 1):
        print(f"   [{i}] {header}")
    
    # Count rows with data
    row_count_table = ws_table.max_row - 1  # Exclude header
    print(f"\n📊 Total Data Rows: {row_count_table}")
    
    # Show first 5 rows as examples
    print(f"\n📝 First 5 Data Rows:")
    for row_idx in range(2, min(7, ws_table.max_row + 1)):
        row_data = []
        for col_idx in range(1, len(headers_table) + 1):
            cell_value = ws_table.cell(row_idx, col_idx).value
            row_data.append(str(cell_value)[:30] if cell_value else "")
        print(f"   Row {row_idx}: {row_data}")

except Exception as e:
    print(f"❌ Error reading Table file: {e}")

print("\n" + "=" * 80)
print("COMPARISON")
print("=" * 80)

try:
    # Check if Complete has additional columns
    complete_cols = set(headers_complete)
    table_cols = set(headers_table)
    
    print(f"\n📊 Structure Comparison:")
    print(f"   Complete.xlsx columns: {len(headers_complete)}")
    print(f"   Table.xlsx columns: {len(headers_table)}")
    
    if complete_cols != table_cols:
        print(f"\n🔍 Different columns found:")
        
        extra_in_complete = complete_cols - table_cols
        if extra_in_complete:
            print(f"\n   ➕ In COMPLETE but not in TABLE:")
            for col in sorted(extra_in_complete):
                print(f"      • {col}")
        
        extra_in_table = table_cols - complete_cols
        if extra_in_table:
            print(f"\n   ➖ In TABLE but not in COMPLETE:")
            for col in sorted(extra_in_table):
                print(f"      • {col}")
    else:
        print(f"\n✅ Same column structure")
    
    print(f"\n📊 Data Volume Comparison:")
    print(f"   Complete.xlsx: {row_count_complete} records")
    print(f"   Table.xlsx: {row_count_table} records")
    
    if row_count_complete > row_count_table:
        print(f"   ✨ Complete has {row_count_complete - row_count_table} MORE records!")
    elif row_count_complete < row_count_table:
        print(f"   Note: Complete has fewer records")
    else:
        print(f"   📌 Same number of records")

except Exception as e:
    print(f"❌ Error in comparison: {e}")

print("\n" + "=" * 80)
print("NEXT STEPS")
print("=" * 80)
print("""
If Complete.xlsx has MORE records:
  → Could improve mapping coverage (more Polaris codes in master data)

If Complete.xlsx has MORE columns:
  → Could have alternative code mappings (e.g., different job code formats)

If Complete.xlsx is IDENTICAL:
  → Already using the best available data
""")
