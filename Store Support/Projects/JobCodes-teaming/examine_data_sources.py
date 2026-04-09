"""
Examine data sources for the proper join strategy:
1. Polaris - authorize source for job codes
2. CoreHR - user/person IDs and roles
3. TMS Data - team/teaming alignment
"""

import openpyxl
from pathlib import Path

print("="*80)
print("EXAMINING DATA SOURCES FOR JOBCODES DASHBOARD")
print("="*80)

# 1. Examine TMS Data (3).xlsx
print("\n" + "="*80)
print("1. TMS DATA (3).xlsx - TEAMS/TEAMING DATA")
print("="*80)

tms_file = Path(r"c:\Users\krush\OneDrive - Walmart Inc\Documents\VSCode\Activity_Hub\Store Support\Projects\JobCodes-teaming\Teaming\TMS Data (3).xlsx")

if tms_file.exists():
    print(f"✅ Found: {tms_file.name}\n")
    
    try:
        wb = openpyxl.load_workbook(tms_file)
        print(f"📊 Sheet Names: {wb.sheetnames}\n")
        
        ws = wb.active
        print(f"📄 Active Sheet: {ws.title}")
        print(f"📈 Dimensions: {ws.dimensions}\n")
        
        # Get headers
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value))
        
        print(f"📋 Columns ({len(headers)}):")
        for i, header in enumerate(headers, 1):
            print(f"   [{i:2d}] {header}")
        
        # Count rows
        row_count = ws.max_row - 1
        print(f"\n📊 Total Data Rows: {row_count}")
        
        # Show first 5 rows as examples
        print(f"\n📝 First 5 Data Rows:")
        for row_idx in range(2, min(7, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, len(headers) + 1):
                cell_value = ws.cell(row_idx, col_idx).value
                cell_str = str(cell_value)[:40] if cell_value else ""
                row_data.append(cell_str)
            print(f"   Row {row_idx}: {row_data[:5]}...")  # Show first 5 columns
        
    except Exception as e:
        print(f"❌ Error reading TMS file: {e}")
else:
    print(f"❌ Not found: {tms_file}")

# 2. Document the data sources we need to query
print("\n" + "="*80)
print("2. DATA SOURCES SUMMARY")
print("="*80)

print("""
POLARIS (Job Code Authority):
📊 Source: polaris-analytics-prod.us_walmart.vw_polaris_current_schedule
   Purpose: Primary job codes, job names, and potentially User IDs
   Action: Need to query - examine what columns available
   
COREHR (User/Person Data):
📊 Source: wmt-corehr-prod.US_HUDI.UNIFIED_PROFILE_SENSITIVE_VW  
   Purpose: User IDs, roles, role types
   Action: Need to query - examine what columns available
   
TEAMS/TEAMING:
📊 Source: TMS Data (3).xlsx (already examined above)
   Purpose: Team/workgroup alignment for Teaming tab
   Action: Map job codes to teams

ENRICHMENT DATA:
📊 Source: Job_Code_Master_Table.xlsx (50 columns)
   Purpose: Category, Job Family, PG Level, etc.
   Action: Already available - full join to Polaris codes
""")

# 3. Proposed join strategy
print("\n" + "="*80)
print("3. PROPOSED FULL JOIN STRATEGY")
print("="*80)

print("""
┌─────────────────────────────────────────────────────────────┐
│ STEP 1: START WITH POLARIS                                  │
│ └─ Load all 271 SMART Job Codes                             │
│    Extract: job_code, job_name, [other columns to identify] │
│                                                              │
│ STEP 2: FULL JOIN + Job_Code_Master_Table.xlsx              │
│ └─ Match on SMART Job Code                                  │
│    Add: Category, Job Family, PG Level, Team, etc.          │
│    If no match: BLANK (team reviews later)                  │
│                                                              │
│ STEP 3: FULL JOIN + TMS_Data (3).xlsx                       │
│ └─ Match on job code                                        │
│    Add: Team/group alignment, teaming details               │
│    If no match: BLANK (marked as "Missing Teaming Data")    │
│                                                              │
│ STEP 4: LOOKUP User ID, Role, Role Type                     │
│ └─ Option A: From Polaris if available                      │
│ └─ Option B: From CoreHR if available                       │
│ └─ Continue filling blanks for team review                  │
│                                                              │
│ RESULT: Complete dataset with:                              │
│ ✓ All 271 Polaris codes (authoritative)                     │
│ ✓ Enrichment data where available                           │
│ ✓ Teaming data where available                              │
│ ✓ Blanks for team to review & update                        │
└─────────────────────────────────────────────────────────────┘
""")

# 4. Next actions
print("\n" + "="*80)
print("4. NEXT ACTIONS NEEDED FROM YOU")
print("="*80)

print("""
❓ QUESTIONS TO ANSWER:

1. Can you confirm the exact columns in Polaris that we should use?
   - Job Code column name?
   - Job Name column name?
   - Does Polaris have User ID columns? If yes, which ones?
   - Does Polaris have Role/Role Type info?
   
2. Can you confirm the exact columns in CoreHR that we should map to?
   - Which column links to job codes (if any)?
   - Which column has User ID?
   - Which column has Role?
   - Which column has Role Type?
   
3. For TMS Data (3).xlsx:
   - Should we match on SMART Job Code directly?
   - Or is there a job code column that needs joining?
   - Which columns represent the "team" data we need for Teaming tab?

✅ ONCE CONFIRMED:
   I will write Python scripts to:
   ✓ Query Polaris and CoreHR data
   ✓ Build the proper full join DataFrames
   ✓ Generate the complete enriched job codes file
   ✓ Mark blanks as "Review Needed" for your team
""")
