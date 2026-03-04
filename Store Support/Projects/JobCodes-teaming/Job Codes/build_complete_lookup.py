#!/usr/bin/env python3
"""
Build a complete job code -> User ID lookup from AMP Roles.xlsx
This will be used to populate missing User IDs and validate existing ones
"""

import openpyxl
import csv
from collections import OrderedDict

print("=" * 80)
print("BUILDING COMPLETE JOB CODE LOOKUP")
print("=" * 80)
print()

# Step 1: Extract existing mappings from AMP Roles.xlsx
print("Step 1: Extracting existing job code -> User ID mappings...")
original_wb = openpyxl.load_workbook('AMP Roles.xlsx')
original_ws = original_wb.active

lookup = OrderedDict()
missing_user_ids_rows = []

for row_idx in range(2, original_ws.max_row + 1):
    job_code = original_ws.cell(row_idx, 3).value      # Column C: Job Code
    user_role = original_ws.cell(row_idx, 4).value     # Column D: User Role/ID
    role = original_ws.cell(row_idx, 2).value          # Column B: Role
    role_type = original_ws.cell(row_idx, 1).value     # Column A: Role Type
    
    job_code_clean = str(job_code).strip() if job_code else ""
    
    if job_code_clean:
        if user_role and str(user_role).strip():
            user_id_clean = str(user_role).strip()
            # Store the mapping with additional context
            lookup[job_code_clean] = {
                'user_id': user_id_clean,
                'role': str(role).strip() if role else '',
                'role_type': str(role_type).strip() if role_type else '',
                'row': row_idx
            }
        else:
            missing_user_ids_rows.append({
                'row': row_idx,
                'job_code': job_code_clean,
                'role': str(role).strip() if role else ''
            })

print(f"  ✓ Found {len(lookup)} job codes with User IDs")
print(f"  ✓ Found {len(missing_user_ids_rows)} job codes with MISSING User IDs")
print()

# Step 2: Show missing job codes
print("Step 2: Identifying missing job codes...")
print()
if missing_user_ids_rows:
    print(f"Missing User IDs for {len(missing_user_ids_rows)} job codes:")
    for item in missing_user_ids_rows[:15]:
        print(f"  Row {item['row']}: {item['job_code']} ({item['role']})")
    if len(missing_user_ids_rows) > 15:
        print(f"  ... and {len(missing_user_ids_rows) - 15} more")
print()

# Step 3: Save to CSV
print("Step 3: Saving lookup to CSV...")
output_file = 'Job_Code_Lookup_Complete.csv'

with open(output_file, 'w', newline='', encoding='utf-8') as f:
    writer = csv.DictWriter(f, fieldnames=['SMART Job Code', 'User ID', 'Role', 'Role Type'])
    writer.writeheader()
    
    for job_code, data in lookup.items():
        writer.writerow({
            'SMART Job Code': job_code,
            'User ID': data['user_id'],
            'Role': data['role'],
            'Role Type': data['role_type']
        })

print(f"  ✓ Saved to: {output_file}")
print(f"  ✓ Total entries: {len(lookup)}")
print()

# Step 4: Create a separate list of job codes needing attention
print("Step 4: Saving missing job codes to separate file...")
missing_file = 'Missing_User_IDs.csv'

with open(missing_file, 'w', newline='', encoding='utf-8') as f:
    writer = csv.DictWriter(f, fieldnames=['row', 'job_code', 'role'])
    writer.writeheader()
    writer.writerows(missing_user_ids_rows)

print(f"  ✓ Saved to: {missing_file}")
print(f"  ✓ Requires {len(missing_user_ids_rows)} lookups")
print()

# Step 5: Show summary statistics
print("=" * 80)
print("SUMMARY")
print("=" * 80)
print(f"Complete mappings: {len(lookup)}")
print(f"Missing mappings: {len(missing_user_ids_rows)}")
print(f"Total job codes: {len(lookup) + len(missing_user_ids_rows)}")
print(f"Coverage: {len(lookup) / (len(lookup) + len(missing_user_ids_rows)) * 100:.1f}%")
print()

# Step 6: Save lookup as Excel for use by create_corrected_final.py
print("Step 6: Creating Excel lookup file for correction script...")
lookup_wb = openpyxl.Workbook()
lookup_ws = lookup_wb.active
lookup_ws.title = "Lookup"

# Add headers
lookup_ws['A1'] = 'SMART Job Code'
lookup_ws['B1'] = 'User ID'
lookup_ws['C1'] = 'Role'
lookup_ws['D1'] = 'Role Type'

# Add data
for row_idx, (job_code, data) in enumerate(lookup.items(), start=2):
    lookup_ws[f'A{row_idx}'] = job_code
    lookup_ws[f'B{row_idx}'] = data['user_id']
    lookup_ws[f'C{row_idx}'] = data['role']
    lookup_ws[f'D{row_idx}'] = data['role_type']

lookup_wb.save('Job_Code_Master_Complete.xlsx')
print(f"  ✓ Created: Job_Code_Master_Complete.xlsx")
print()

print("=" * 80)
print("NEXT STEPS:")
print("=" * 80)
print("1. Review Missing_User_IDs.csv to identify patterns")
print("2. Use BigQuery or other data sources to populate missing User IDs")
print("3. Update Job_Code_Master_Complete.xlsx with missing mappings")
print("4. Run: python create_corrected_final.py")
print("=" * 80)
