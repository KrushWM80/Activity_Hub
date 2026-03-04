import openpyxl
import os

# The lookup data with job codes and their correct User IDs
lookup = {
    '1-202-2104': 'n0j0492.s03348',
    '1-203-2003': 'a0f0m4u.s01265',
    '1-27-237': 's0h0vrm.s02862',
    '1-46-501': 'z0h06qo.s03342',
    '1-580-131': 'rlm002c.s02884',
    '1-930-1471': 'c0f0egf.s01105',
    '59-645-7310': 'j0m0dtd.s01089'
}

print("=" * 60)
print("CREATING CORRECTED FILE WITH OPENPYXL")
print("=" * 60)
print()
print("Loaded lookup data:")
print(f"  Total codes: {len(lookup)}")
print()

# Load original file
print("Loading original AMP Roles.xlsx...")
wb = openpyxl.load_workbook('AMP Roles.xlsx')
ws = wb.active
print(f"  Loaded sheet: {ws.title}")
print(f"  Total rows: {ws.max_row}")
print()

# Find the job code and user role columns
header_row = ws[1]
job_code_col = None
user_role_col = None

for col_idx, cell in enumerate(header_row, 1):
    if cell.value and 'Job Code' in str(cell.value):
        job_code_col = col_idx
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        print(f"  Found Job Code column at: {col_letter}")
    if cell.value and ('User Role' in str(cell.value) or 'User ID' in str(cell.value)):
        user_role_col = col_idx
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        print(f"  Found User Role column at: {col_letter}")

print()

if not job_code_col or not user_role_col:
    print("ERROR: Could not find Job Code or User Role columns")
    exit(1)

# Update User IDs
updates = 0
missing_count = 0
matched_count = 0

for row_idx in range(2, ws.max_row + 1):
    job_code_cell = ws.cell(row_idx, job_code_col)
    user_role_cell = ws.cell(row_idx, user_role_col)
    
    job_code = job_code_cell.value
    
    if job_code:
        if job_code in lookup:
            correct_user_id = lookup[job_code]
            if user_role_cell.value != correct_user_id:
                user_role_cell.value = correct_user_id
                updates += 1
            matched_count += 1
        else:
            if not user_role_cell.value:
                missing_count += 1

print(f"Processing Results:")
print(f"  Matched job codes: {matched_count}")
print(f"  Updated User IDs: {updates}")
print(f"  Missing User IDs: {missing_count}")
print()

# Save as new file
output_file = 'AMP Roles_CORRECTED.xlsx'
print(f"Saving to: {output_file}")
wb.save(output_file)

# Check file size
file_size = os.path.getsize(output_file)
print(f"  File size: {file_size:,} bytes")
print()

# Verify
print("Verifying file...")
wb2 = openpyxl.load_workbook(output_file)
ws2 = wb2.active
print(f"  Rows in new file: {ws2.max_row}")
print(f"  File is readable: YES ✓")
print()

# Show sample corrections
print("Sample of corrected entries:")
for row_idx in range(2, min(10, ws2.max_row + 1)):
    job_code = ws2.cell(row_idx, job_code_col).value
    user_id = ws2.cell(row_idx, user_role_col).value
    if job_code:
        print(f"  Row {row_idx}: {job_code} → {user_id}")

print()
print("=" * 60)
print("✓ FILE CREATED SUCCESSFULLY")
print("=" * 60)
