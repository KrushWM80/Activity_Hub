#!/usr/bin/env python3
"""
Query BigQuery CoreHR dataset to get User IDs for missing job codes
"""

import csv
import os
from google.cloud import bigquery
import openpyxl

print("=" * 90)
print("BIGQUERY QUERY - MISSING USER IDS")
print("=" * 90)
print()

# Step 1: Load missing job codes
print("Step 1: Loading missing job codes...")
missing_codes = []

with open('Missing_User_IDs.csv', 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        missing_codes.append(row['job_code'].strip())

print(f"  ✓ Loaded {len(missing_codes)} missing job codes")
print(f"    Examples: {missing_codes[:5]}")
print()

# Step 2: Initialize BigQuery client
print("Step 2: Connecting to BigQuery...")
try:
    # Create client without restricting to one project (allows cross-project queries)
    client = bigquery.Client()
    print(f"  ✓ Connected to BigQuery (default project)")
except Exception as e:
    print(f"  ✗ Error: {e}")
    print()
    print("Make sure:")
    print("  1. gcloud is installed and configured")
    print("  2. Run: gcloud auth application-default login")
    print("  3. Set GCP project: gcloud config set project wmt-assetprotection-prod")
    exit(1)

print()

# Step 3: Query for missing job codes (using Polaris dataset primarily)
print("Step 3: Querying BigQuery for User IDs...")
print()

# Try multiple possible table names and structures
possible_queries = [
    # Query 1: Polaris - has worker_id (User ID) and job_code (SMART code)
    """
    SELECT DISTINCT
        job_code,
        worker_id as user_id,
        first_name,
        last_name,
        job_nm
    FROM `polaris-analytics-prod.us_walmart.vw_polaris_current_schedule`
    WHERE job_code IN ({job_codes})
    LIMIT 1000
    """,
    
    # Query 2: Try CoreHR unified profile with job code
    """
    SELECT DISTINCT
        job_code,
        associate_id as user_id,
        first_name,
        last_name,
        job_title
    FROM `wmt-corehr-prod.US_HUDI.UNIFIED_PROFILE_SENSITIVE_VW`
    WHERE job_code IN ({job_codes})
    LIMIT 1000
    """,
    
    # Query 3: Try broader search across tables
    """
    SELECT DISTINCT
        job_code,
        worker_id as user_id,
        first_name,
        last_name,
        job_nm
    FROM `polaris-analytics-prod.us_walmart.vw_polaris_employee_master`
    WHERE job_code IN ({job_codes})
    LIMIT 1000
    """,
]

# Format job codes for SQL IN clause
job_codes_sql = ', '.join([f"'{code}'" for code in missing_codes])

found_results = False
results = {}

for idx, query_template in enumerate(possible_queries, 1):
    if found_results:
        break
    
    try:
        query = query_template.format(job_codes=job_codes_sql)
        print(f"  Attempt {idx}: Querying BigQuery...")
        
        query_job = client.query(query)
        rows = list(query_job.result())
        
        if rows:
            print(f"  ✓ Query succeeded! Found {len(rows)} results")
            print()
            print("    Sample results:")
            for row in rows[:5]:
                print(f"      {dict(row)}")
            print()
            
            # Store results
            for row in rows:
                job_code = row.get('job_code')
                user_id = row.get('user_id')
                if job_code and user_id:
                    results[job_code.strip()] = user_id.strip()
            
            found_results = True
            break
    except Exception as e:
        error_str = str(e).lower()
        if 'not found' in error_str or 'does not exist' in error_str:
            print(f"  ✗ Table/column not found (trying next...)")
        else:
            print(f"  ✗ Error: {e}")

if not found_results:
    print("  ⚠ Could not find User IDs in BigQuery")
    print()
    print("Next steps:")
    print("  1. Check the actual table/column names in your BigQuery dataset")
    print("  2. Update the query_template in this script")
    print("  3. Re-run this script")
    print()
    print("To explore tables, run:")
    print("  bq ls --project_id=wmt-assetprotection-prod Store_Support_Dev")
    print("  bq show --project_id=wmt-assetprotection-prod Store_Support_Dev.TABLE_NAME")
    exit(1)

print()

# Step 4: Load existing lookup
print("Step 4: Loading existing lookup data...")
existing_lookup = {}

try:
    wb = openpyxl.load_workbook('Job_Code_Master_Complete.xlsx')
    ws = wb.active
    
    for row_idx in range(2, ws.max_row + 1):
        job_code = ws.cell(row_idx, 1).value
        user_id = ws.cell(row_idx, 2).value
        
        if job_code and user_id:
            existing_lookup[str(job_code).strip()] = str(user_id).strip()
    
    print(f"  ✓ Loaded {len(existing_lookup)} existing mappings")
except Exception as e:
    print(f"  ✗ Could not load existing lookup: {e}")

print()

# Step 5: Merge and update
print("Step 5: Merging new results with existing data...")

for job_code in missing_codes:
    if job_code in results:
        existing_lookup[job_code] = results[job_code]

print(f"  ✓ Total mappings now: {len(existing_lookup)}")
print()

# Step 6: Save updated lookup
print("Step 6: Saving updated lookup file...")

# Save to Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Lookup"

# Headers
ws['A1'] = 'SMART Job Code'
ws['B1'] = 'User ID'
ws['C1'] = 'Role'
ws['D1'] = 'Role Type'

# Data
for row_idx, (job_code, user_id) in enumerate(sorted(existing_lookup.items()), start=2):
    ws[f'A{row_idx}'] = job_code
    ws[f'B{row_idx}'] = user_id

wb.save('Job_Code_Master_Complete.xlsx')
print(f"  ✓ Updated: Job_Code_Master_Complete.xlsx")

# Save to CSV
with open('Job_Code_Lookup_Complete.csv', 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerow(['SMART Job Code', 'User ID'])
    for job_code, user_id in sorted(existing_lookup.items()):
        writer.writerow([job_code, user_id])

print(f"  ✓ Updated: Job_Code_Lookup_Complete.csv")
print()

# Step 7: Summary
print("=" * 90)
print("SUMMARY")
print("=" * 90)
print(f"Missing job codes checked: {len(missing_codes)}")
print(f"User IDs found from BigQuery: {len(results)}")
print(f"Total mappings in updated lookup: {len(existing_lookup)}")
print()

if len(results) > 0:
    print(f"✓ SUCCESS - Found {len(results)} User IDs in BigQuery!")
    print()
    print("Next steps:")
    print("  1. Run: python create_corrected_final.py")
    print("  2. It will use the updated lookup to populate remaining User IDs")
    print()
else:
    print("⚠ No User IDs found in BigQuery")
    print()
    print("Please verify:")
    print("  1. The table structure (run: bq show Store_Support_Dev.TABLE_NAME)")
    print("  2. The column names for job codes and user IDs")
    print("  3. Update this script with correct table/column names")
    print("  4. Re-run this script")

print("=" * 90)
