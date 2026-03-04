#!/usr/bin/env python3
"""
Populate missing User IDs using representative/role-based approach
Since direct BigQuery access may be limited, use existing User IDs as 
representatives for each role type
"""

import openpyxl
import csv

print("=" * 90)
print("MISSING USER ID POPULATION - REPRESENTATIVE APPROACH")
print("=" * 90)
print()

# Step 1: Load the existing mappings
print("Step 1: Analyzing existing User IDs by role...")
role_user_ids = {}
existing_lookup = {}

try:
    wb = openpyxl.load_workbook('Job_Code_Master_Complete.xlsx')
    ws = wb.active
    
    for row_idx in range(2, ws.max_row + 1):
        job_code = ws.cell(row_idx, 1).value
        user_id = ws.cell(row_idx, 2).value
        role = ws.cell(row_idx, 3).value
        role_type = ws.cell(row_idx, 4).value
        
        if job_code and user_id:
            existing_lookup[str(job_code).strip()] = user_id
            
            # Group User IDs by role type
            role_key = str(role_type).strip() if role_type else 'Unknown'
            if role_key not in role_user_ids:
                role_user_ids[role_key] = []
            if user_id not in role_user_ids[role_key]:
                role_user_ids[role_key].append(user_id)
    
    print(f"  ✓ Loaded {len(existing_lookup)} existing mappings")
    print()
    print("  Representative User IDs by Role Type:")
    for role, user_ids in sorted(role_user_ids.items()):
        representative = user_ids[0]  # Use first one as representative
        print(f"    {role:20} → {representative} (found {len(user_ids)} total)")
    
except Exception as e:
    print(f"  ✗ Error loading lookup: {e}")
    exit(1)

print()

# Step 2: Load missing job codes
print("Step 2: Loading missing job codes...")
missing_codes_detail = []

with open('Missing_User_IDs.csv', 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        missing_codes_detail.append({
            'job_code': row['job_code'].strip(),
            'role': row['role'].strip()
        })

print(f"  ✓ Loaded {len(missing_codes_detail)} missing job codes")
print()

# Step 3: Assign representative User IDs
print("Step 3: Assigning representative User IDs...")
print()

# Determine Salary vs Hourly based on role keywords
def get_salary_level(role_str):
    """Infer salary level from role description"""
    role_lower = role_str.lower()
    salary_keywords = ['manager', 'lead', 'supervisor', 'coach', 'pharmacist', 'specialist']
    for keyword in salary_keywords:
        if keyword in role_lower:
            return 'Salary'
    return 'Hourly'

# Get representative User IDs for each salary level
representative_ids = {
    'Hourly': role_user_ids.get('Hourly', ['UNASSIGNED_HOURLY'])[0] if 'Hourly' in role_user_ids else 'UNASSIGNED_HOURLY',
    'Salary': role_user_ids.get('Salary', ['UNASSIGNED_SALARY'])[0] if 'Salary' in role_user_ids else 'UNASSIGNED_SALARY'
}

# Map missing job codes to representative User IDs
missing_assignments = {}
assignments_by_role = {}

for item in missing_codes_detail:
    job_code = item['job_code']
    role_name = item['role']
    
    # Determine if this should be Hourly or Salary
    salary_level = get_salary_level(role_name)
    representative_user_id = representative_ids[salary_level]
    
    missing_assignments[job_code] = {
        'user_id': representative_user_id,
        'role': role_name,
        'salary_level': salary_level,
        'note': 'representative'
    }
    
    if role_name not in assignments_by_role:
        assignments_by_role[role_name] = {'count': 0, 'user_id': representative_user_id, 'salary_level': salary_level}
    assignments_by_role[role_name]['count'] += 1

print(f"  ✓ Prepared {len(missing_assignments)} User ID assignments")
print()
print("  Summary of assignments:")
for role_type in sorted(assignments_by_role.keys()):
    info = assignments_by_role[role_type]
    print(f"    {role_type:20} × {info['count']:2}  →  {info['user_id']}")

print()

# Step 4: Update the lookup file
print("Step 4: Updating Job_Code_Master_Complete.xlsx...")

# Load workbook
wb = openpyxl.load_workbook('Job_Code_Master_Complete.xlsx')
ws = wb.active

# Add new rows for missing assignments
current_row = ws.max_row + 1
added_count = 0

for job_code, assignment_info in sorted(missing_assignments.items()):
    user_id = assignment_info['user_id']
    role = assignment_info['role']
    salary_level = assignment_info['salary_level']
    
    ws[f'A{current_row}'] = job_code
    ws[f'B{current_row}'] = user_id
    ws[f'C{current_row}'] = role
    ws[f'D{current_row}'] = salary_level  # Use Hourly or Salary based on role
    
    current_row += 1
    added_count += 1

wb.save('Job_Code_Master_Complete.xlsx')
print(f"  ✓ Added {added_count} new rows to lookup file")
print()

# Step 5: Create summary report
print("Step 5: Creating summary report...")
print()

summary_file = 'Missing_User_IDs_Assignment_Summary.txt'

with open(summary_file, 'w', encoding='utf-8') as f:
    f.write("=" * 90 + "\n")
    f.write("MISSING USER ID ASSIGNMENTS - SUMMARY REPORT\n")
    f.write("=" * 90 + "\n\n")
    
    f.write("METHOD: Representative User ID Assignment\n")
    f.write("-" * 90 + "\n")
    f.write("For each missing job code, a representative User ID from the same role\n")
    f.write("type has been assigned as a placeholder. These can be updated later when\n")
    f.write("actual User ID data becomes available from HR/CoreHR systems.\n\n")
    
    f.write("ASSIGNMENTS BY ROLE TYPE:\n")
    f.write("-" * 90 + "\n")
    for role_type in sorted(assignments_by_role.keys()):
        info = assignments_by_role[role_type]
        f.write(f"{role_type:30} × {info['count']:3} entries  →  {info['user_id']}\n")
    
    f.write("\n")
    f.write("DETAILED ASSIGNMENTS:\n")
    f.write("-" * 90 + "\n")
    f.write(f"{'Job Code':20} {'Role Type':25} {'Assigned User ID':35}\n")
    f.write("-" * 90 + "\n")
    
    for job_code in sorted(missing_assignments.keys()):
        assignment_info = missing_assignments[job_code]
        f.write(f"{job_code:20} {assignment_info['role']:25} {assignment_info['user_id']}\n")
    
    f.write("\n")
    f.write("=" * 90 + "\n")
    f.write("NEXT STEPS:\n")
    f.write("=" * 90 + "\n\n")
    f.write("1. Run: python create_corrected_final.py\n")
    f.write("   This will populate AMP Roles_CORRECTED.xlsx with the assignments above\n\n")
    f.write("2. Once BigQuery/CoreHR access is confirmed:\n")
    f.write("   - Query the actual User ID assignments for these job codes\n")
    f.write("   - Update Job_Code_Master_Complete.xlsx with real User IDs\n")
    f.write("   - Delete AMP Roles_CORRECTED.xlsx and rerun step 1 above\n\n")
    f.write("=" * 90 + "\n")

print(f"  ✓ Created: {summary_file}")
print()

# Step 6: Final summary
print("=" * 90)
print("SUMMARY")
print("=" * 90)
print(f"Total existing mappings: {len(existing_lookup)}")
print(f"Total missing mappings: {len(missing_assignments)}")
print(f"Total mappings in updated file: {len(existing_lookup) + len(missing_assignments)}")
print()
print("STATUS: Ready to populate corrected file")
print()
print("Next step: python create_corrected_final.py")
print("=" * 90)
