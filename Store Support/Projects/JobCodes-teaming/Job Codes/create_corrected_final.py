import openpyxl
import os

print("=" * 70)
print("CREATING CORRECTED AMP ROLES FILE")
print("=" * 70)
print()

# Step 1: Load lookup data from Job_Code_Master_Complete.xlsx (or try alternatives)
print("Step 1: Loading job code lookup data...")
try:
    # Try the complete lookup file first
    master_wb = openpyxl.load_workbook('Job_Code_Master_Complete.xlsx')
    master_ws = master_wb.active
    
    lookup = {}
    for row_idx in range(2, master_ws.max_row + 1):
        job_code = master_ws.cell(row_idx, 1).value  # First column
        user_id = master_ws.cell(row_idx, 2).value   # Second column
        
        if job_code and user_id:
            lookup[str(job_code).strip()] = str(user_id).strip()
    
    print(f"  ✓ Loaded {len(lookup)} job code mappings from Job_Code_Master_Complete.xlsx")
    print(f"    Sample: {list(lookup.items())[:3]}")
except Exception as e:
    print(f"  ✗ Error reading Job_Code_Master_Complete: {e}")
    print()
    print("Attempting with Job_Code_Master_Table.xlsx...")
    try:
        master_wb = openpyxl.load_workbook('Job_Code_Master_Table.xlsx')
        master_ws = master_wb.active
        
        lookup = {}
        for row_idx in range(2, master_ws.max_row + 1):
            job_code = master_ws.cell(row_idx, 1).value  # First column
            user_id = master_ws.cell(row_idx, 2).value   # Second column
            
            if job_code and user_id:
                lookup[str(job_code).strip()] = str(user_id).strip()
        
        print(f"  ✓ Loaded {len(lookup)} job code mappings from Job_Code_Master_Table.xlsx")
    except Exception as e2:
        print(f"  ✗ Error reading Job_Code_Master_Table: {e2}")
        print()
        print("Attempting with AMP Roles Updated.xlsx as fallback...")
    try:
        updated_wb = openpyxl.load_workbook('AMP Roles Updated.xlsx')
        updated_ws = updated_wb.active
        
        lookup = {}
        for row_idx in range(2, updated_ws.max_row + 1):
            job_code = updated_ws.cell(row_idx, 3).value  # Column C (Job Code)
            user_id = updated_ws.cell(row_idx, 4).value   # Column D (User ID)
            
            if job_code and user_id:
                lookup[str(job_code).strip()] = str(user_id).strip()
        
        print(f"  ✓ Loaded {len(lookup)} job code mappings from Updated file")
    except Exception as e2:
        print(f"  ✗ Both sources failed: {e2}")
        exit(1)

print()

# Step 2: Load the original AMP Roles.xlsx
print("Step 2: Loading original AMP Roles.xlsx...")
try:
    original_wb = openpyxl.load_workbook('AMP Roles.xlsx')
    original_ws = original_wb.active
    print(f"  ✓ Loaded {original_ws.title}")
    print(f"  ✓ Total rows: {original_ws.max_row}")
except Exception as e:
    print(f"  ✗ Error: {e}")
    exit(1)

print()

# Step 3: Find the correct columns
print("Step 3: Finding job code and user role columns...")
header = original_ws[1]
job_code_col = None
user_role_col = None

for col_idx, cell in enumerate(header, 1):
    if cell.value:
        val = str(cell.value).lower()
        if 'job' in val and 'code' in val:
            job_code_col = col_idx
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            print(f"  ✓ Job Code found at column {col_letter}")
        if 'user' in val and ('role' in val or 'id' in val):
            user_role_col = col_idx
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            print(f"  ✓ User Role found at column {col_letter}")

if not job_code_col or not user_role_col:
    print(f"  ✗ ERROR: Could not locate columns")
    print(f"    Job Code column: {job_code_col}")
    print(f"    User Role column: {user_role_col}")
    exit(1)

print()

# Step 4: Update User IDs
print("Step 4: Updating User IDs...")
updates = 0
filled = 0

for row_idx in range(2, original_ws.max_row + 1):
    job_code_cell = original_ws.cell(row_idx, job_code_col)
    user_role_cell = original_ws.cell(row_idx, user_role_col)
    
    job_code = job_code_cell.value
    
    if job_code:
        job_code_str = str(job_code).strip()
        
        if job_code_str in lookup:
            correct_user_id = lookup[job_code_str]
            current_user_id = user_role_cell.value
            
            if current_user_id != correct_user_id:
                user_role_cell.value = correct_user_id
                if current_user_id:
                    updates += 1
                else:
                    filled += 1

print(f"  ✓ Updated incorrect User IDs: {updates}")
print(f"  ✓ Filled missing User IDs: {filled}")
print(f"  ✓ Total changes: {updates + filled}")
print()

# Step 5: Save corrected file
output_file = 'AMP Roles_CORRECTED.xlsx'
print(f"Step 5: Saving corrected file...")
try:
    original_wb.save(output_file)
    file_size = os.path.getsize(output_file)
    print(f"  ✓ Saved: {output_file}")
    print(f"  ✓ File size: {file_size:,} bytes")
except Exception as e:
    print(f"  ✗ Error: {e}")
    exit(1)

print()

# Step 6: Verify
print("Step 6: Verifying file integrity...")
try:
    verify_wb = openpyxl.load_workbook(output_file)
    verify_ws = verify_wb.active
    print(f"  ✓ File is readable")
    print(f"  ✓ Total rows: {verify_ws.max_row}")
    
    # Show sample entries
    print()
    print("Sample corrected entries:")
    count = 0
    for row_idx in range(2, verify_ws.max_row + 1):
        job_code = verify_ws.cell(row_idx, job_code_col).value
        user_id = verify_ws.cell(row_idx, user_role_col).value
        if job_code and user_id and count < 5:
            print(f"  Row {row_idx}: {job_code} → {user_id}")
            count += 1
except Exception as e:
    print(f"  ✗ Verification failed: {e}")
    exit(1)

print()
print("=" * 70)
print("✅ SUCCESS - File created and verified")
print("=" * 70)
print()
print(f"Location: {os.path.abspath(output_file)}")
print()
print("The file is now ready to use. Open it in Excel to view the data.")
