"""
Build automatic job code mapping from Excel master data
Maps Polaris SMART codes to Excel enrichment data using two strategies:
1. Workday Job Code matching (90 codes)
2. Job Title matching (19 codes)
"""

import json
import sqlite3
from pathlib import Path
from openpyxl import load_workbook

# Load Polaris codes from database
polaris_db = Path(r"c:\Users\krush\OneDrive - Walmart Inc\Documents\VSCode\Activity_Hub\Store Support\Projects\JobCodes-teaming\Teaming\dashboard\backend\cache\jobcodes_cache.db")

conn = sqlite3.connect(polaris_db)
cursor = conn.cursor()

# Get Polaris codes
cursor.execute("SELECT job_code, job_nm FROM polaris_job_codes")
polaris_codes = {}
for smart_code, job_name_with_code in cursor.fetchall():
    # Extract clean job name (remove the embedded code)
    clean_name = job_name_with_code.replace(f" {smart_code}", "").strip()
    polaris_codes[smart_code] = {
        "job_nm": job_name_with_code,
        "clean_name": clean_name
    }

print(f"✅ Loaded {len(polaris_codes)} Polaris codes")

# Load Excel master data
excel_file = Path(r"c:\Users\krush\OneDrive - Walmart Inc\Documents\VSCode\Activity_Hub\Store Support\Projects\JobCodes-teaming\Job Codes\Job_Code_Master_Table.xlsx")
wb = load_workbook(excel_file)
ws = wb.active

# Parse Excel data
excel_data = {}
excel_by_workday = {}
excel_by_title = {}

# Get headers
headers = [cell.value for cell in ws[1]]
col_map = {h: i for i, h in enumerate(headers)}

print(f"📊 Excel columns found: {len(headers)}")

for row_idx in range(2, ws.max_row + 1):
    row_values = [ws.cell(row_idx, col_idx).value for col_idx in range(1, len(headers) + 1)]
    
    smart_code = row_values[col_map.get("SMART Job Code")]
    workday_code = row_values[col_map.get("Workday Job Code")]
    job_title = row_values[col_map.get("Job Title")]
    category = row_values[col_map.get("Category")]
    job_family = row_values[col_map.get("Job Family")]
    pg_level = row_values[col_map.get("PG Level")]
    team = row_values[col_map.get("Team")]
    workgroup = row_values[col_map.get("Workgroup")]
    supervisor = row_values[col_map.get("Supervisor?")]
    reports_to = row_values[col_map.get("Reports to Title")]
    
    if smart_code:
        excel_data[smart_code] = {
            "workday_code": workday_code,
            "job_title": job_title,
            "category": category,
            "job_family": job_family,
            "pg_level": pg_level,
            "team": team,
            "workgroup": workgroup,
            "supervisor": supervisor,
            "reports_to": reports_to
        }
        
        # Index by Workday code for strategy 1
        if workday_code:
            if workday_code not in excel_by_workday:
                excel_by_workday[workday_code] = []
            excel_by_workday[workday_code].append(smart_code)
        
        # Index by title for strategy 2 (clean titles)
        if job_title:
            clean_title = job_title.strip()
            if clean_title not in excel_by_title:
                excel_by_title[clean_title] = []
            excel_by_title[clean_title].append(smart_code)

print(f"✅ Loaded {len(excel_data)} Excel job codes")
print(f"✅ Indexed {len(excel_by_workday)} unique Workday codes")
print(f"✅ Indexed {len(excel_by_title)} unique job titles")

# ============================================================================
# STRATEGY 1: Match Polaris code to Excel "Workday Job Code" column
# ============================================================================
print("\n" + "="*80)
print("STRATEGY 1: Workday Job Code Matching")
print("="*80)

strategy1_matches = {}
strategy1_details = []

for smart_code, polaris_info in polaris_codes.items():
    if smart_code in excel_by_workday:
        excel_smart_code = excel_by_workday[smart_code][0]
        enrichment = excel_data[excel_smart_code]
        strategy1_matches[smart_code] = enrichment
        strategy1_details.append({
            "polaris_smart_code": smart_code,
            "polaris_job_name": polaris_info["job_nm"],
            "match_type": "Workday Code",
            "enrichment": enrichment
        })

print(f"✅ Found {len(strategy1_matches)} direct Workday code matches")
print(f"   Examples:")
for detail in strategy1_details[:5]:
    print(f"     • {detail['polaris_smart_code']} ({detail['polaris_job_name'][:40]})")
    print(f"       → Category: {detail['enrichment'].get('category')}, Job Family: {detail['enrichment'].get('job_family')}")

# ============================================================================
# STRATEGY 2: Match by clean job title
# ============================================================================
print("\n" + "="*80)
print("STRATEGY 2: Job Title Matching")
print("="*80)

strategy2_matches = {}
strategy2_details = []

for smart_code, polaris_info in polaris_codes.items():
    if smart_code not in strategy1_matches:  # Don't double-count
        clean_polaris_name = polaris_info["clean_name"]
        
        if clean_polaris_name in excel_by_title:
            excel_smart_code = excel_by_title[clean_polaris_name][0]
            enrichment = excel_data[excel_smart_code]
            strategy2_matches[smart_code] = enrichment
            strategy2_details.append({
                "polaris_smart_code": smart_code,
                "polaris_job_name": polaris_info["job_nm"],
                "match_type": "Title Match",
                "enrichment": enrichment
            })

print(f"✅ Found {len(strategy2_matches)} additional title-based matches")
print(f"   Examples:")
for detail in strategy2_details[:5]:
    print(f"     • {detail['polaris_smart_code']} ({detail['polaris_job_name'][:40]})")
    print(f"       → Category: {detail['enrichment'].get('category')}, Job Family: {detail['enrichment'].get('job_family')}")

# ============================================================================
# COMBINED RESULTS
# ============================================================================
print("\n" + "="*80)
print("COMBINED MAPPING RESULTS")
print("="*80)

total_mapped = len(strategy1_matches) + len(strategy2_matches)
coverage = (total_mapped / len(polaris_codes)) * 100

print(f"""
┌─────────────────────────────────┐
│ MAPPING SUMMARY                 │
├─────────────────────────────────┤
│ Workday Code Matches:  {len(strategy1_matches):3d} / 271 │
│ Title-Based Matches:   {len(strategy2_matches):3d} / 271 │
│ Total Mapped:          {total_mapped:3d} / 271 │
│ Coverage:              {coverage:5.1f}%      │
│ Unmapped:              {len(polaris_codes) - total_mapped:3d} / 271 │
└─────────────────────────────────┘
""")

# ============================================================================
# Create mapping CSV for database sync
# ============================================================================
print("\n" + "="*80)
print("CREATING MAPPING CSV")
print("="*80)

mapping_csv = []
mapping_csv.append("smart_code,category,job_family,pg_level,team,workgroup,supervisor,reports_to")

all_mapped = {**strategy1_matches, **strategy2_matches}

for smart_code in sorted(all_mapped.keys()):
    enrichment = all_mapped[smart_code]
    csv_line = ",".join([
        f'"{smart_code}"',
        f'"{enrichment.get("category", "")}"',
        f'"{enrichment.get("job_family", "")}"',
        f'"{enrichment.get("pg_level", "")}"',
        f'"{enrichment.get("team", "")}"',
        f'"{enrichment.get("workgroup", "")}"',
        f'"{enrichment.get("supervisor", "")}"',
        f'"{enrichment.get("reports_to", "")}"',
    ])
    mapping_csv.append(csv_line)

# Save mapping CSV
mapping_file = Path(r"c:\Users\krush\OneDrive - Walmart Inc\Documents\VSCode\Activity_Hub\Store Support\Projects\JobCodes-teaming\job_code_enrichment_mapping.csv")
mapping_file.write_text("\n".join(mapping_csv))

print(f"✅ Saved mapping file: {mapping_file.name}")
print(f"   Contains {len(mapping_csv) - 1} enriched job codes")

# ============================================================================
# Create JSON version for reference
# ============================================================================
mapping_json = {}
for smart_code in sorted(all_mapped.keys()):
    mapping_json[smart_code] = all_mapped[smart_code]

mapping_json_file = Path(r"c:\Users\krush\OneDrive - Walmart Inc\Documents\VSCode\Activity_Hub\Store Support\Projects\JobCodes-teaming\job_code_enrichment_mapping.json")
with open(mapping_json_file, 'w') as f:
    json.dump(mapping_json, f, indent=2)

print(f"✅ Saved JSON mapping: {mapping_json_file.name}")

# ============================================================================
# Show unmapped codes
# ============================================================================
print("\n" + "="*80)
print("UNMAPPED POLARIS CODES (Will need manual mapping)")
print("="*80)

unmapped = []
for smart_code in sorted(polaris_codes.keys()):
    if smart_code not in all_mapped:
        unmapped.append(smart_code)

print(f"Total unmapped: {len(unmapped)}")
print(f"\nFirst 20 unmapped codes:")
for code in unmapped[:20]:
    print(f"  • {code} - {polaris_codes[code]['job_nm']}")

if len(unmapped) > 20:
    print(f"  ... and {len(unmapped) - 20} more")

conn.close()
print("\n✅ Analysis complete!")
