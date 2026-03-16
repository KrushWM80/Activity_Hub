"""
Create AMP Roles Updated.xlsx - Fresh Excel file with all job code mappings
Uses openpyxl with proper Excel formatting
"""
import sys
import os

# Try openpyxl first
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    use_openpyxl = True
    print("✓ Using openpyxl library")
except ImportError:
    print("⚠ openpyxl not available, building raw Excel structure")
    use_openpyxl = False

import pandas as pd
from pathlib import Path

def create_with_openpyxl():
    """Create Excel file using openpyxl (best method)"""
    print("\n" + "="*70)
    print("CREATING AMP ROLES UPDATED.XLSX WITH OPENPYXL")
    print("="*70)
    
    # Load the lookup data
    lookup_path = Path(r"C:\Users\krush\Downloads\SMART_JobCode_Lookup.csv")
    if not lookup_path.exists():
        print(f"❌ Lookup file not found: {lookup_path}")
        return False
    
    lookup_df = pd.read_csv(lookup_path)
    print(f"\n✓ Loaded {len(lookup_df)} job codes with mappings")
    print(f"  Columns: {list(lookup_df.columns)}")
    
    # Load template to understand structure
    template_path = Path(r"C:\Users\krush\OneDrive - Walmart Inc\Documents\VSCode\Activity_Hub\Store Support\Projects\JobCodes-teaming\Job Codes\AMP Roles.xlsx.backup_before_update")
    
    if template_path.exists():
        template_df = pd.read_excel(template_path)
        print(f"✓ Loaded template with {len(template_df)} rows")
        print(f"  Template columns: {list(template_df.columns)}")
    else:
        print("⚠ Template not found, creating from scratch")
        template_df = None
    
    # Create new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "AMP Roles"
    
    print("\n1. Creating sheet structure...")
    
    # Define headers based on original
    if template_df is not None:
        headers = list(template_df.columns)
    else:
        headers = ['Current Job Code', 'Role', 'User ID', 'Status']
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    print(f"✓ Headers written: {headers}")
    
    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    if len(headers) > 3:
        ws.column_dimensions['D'].width = 12
    
    print("\n2. Processing and inserting data...")
    
    # Create lookup dictionary for quick access
    lookup_dict = {}
    for _, row in lookup_df.iterrows():
        smart_code = str(row['SMART Job Code']).strip()
        lookup_dict[smart_code] = {
            'Role Type': row.get('Role Type', ''),
            'User ID': row.get('User ID', '')
        }
    
    print(f"✓ Created lookup dictionary with {len(lookup_dict)} entries")
    
    # Load template data if available
    if template_df is not None:
        # Write all rows from template, updating with lookup data where available
        for row_num, (_, row) in enumerate(template_df.iterrows(), 2):
            job_code = str(row[headers[0]]).strip() if headers[0] in row else ""
            
            ws.cell(row=row_num, column=1).value = job_code
            
            # Check if this job code is in our lookup
            if job_code in lookup_dict:
                # Write Role Type
                ws.cell(row=row_num, column=2).value = lookup_dict[job_code]['Role Type']
                # Write User ID
                ws.cell(row=row_num, column=3).value = lookup_dict[job_code]['User ID']
                ws.cell(row=row_num, column=2).fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
                ws.cell(row=row_num, column=3).fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
            else:
                # Copy remaining data from template
                for col_num in range(2, len(headers) + 1):
                    if headers[col_num - 1] in row:
                        ws.cell(row=row_num, column=col_num).value = row[headers[col_num - 1]]
        
        print(f"✓ Written {len(template_df)} rows with update highlighting")
    
    # Save the workbook
    output_path = Path(r"C:\Users\krush\OneDrive - Walmart Inc\Documents\VSCode\Activity_Hub\Store Support\Projects\JobCodes-teaming\Job Codes\AMP Roles Updated.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    wb.save(output_path)
    print(f"\n✓ File saved: {output_path}")
    
    # Verify
    file_size = output_path.stat().st_size
    print(f"✓ File size: {file_size:,} bytes")
    
    # Verify by reading back
    verify_df = pd.read_excel(output_path)
    print(f"✓ Verification: {len(verify_df)} rows, {len(verify_df.columns)} columns")
    
    matched_count = verify_df[verify_df['User ID'].notna()].shape[0]
    print(f"✓ Rows with User ID: {matched_count}")
    
    return True

def create_with_pandas():
    """Fallback: Create using pandas.ExcelWriter"""
    print("\n" + "="*70)
    print("CREATING AMP ROLES UPDATED.XLSX WITH PANDAS")
    print("="*70)
    
    # Load the lookup data
    lookup_path = Path(r"C:\Users\krush\Downloads\SMART_JobCode_Lookup.csv")
    if not lookup_path.exists():
        print(f"❌ Lookup file not found: {lookup_path}")
        return False
    
    lookup_df = pd.read_csv(lookup_path)
    print(f"\n✓ Loaded {len(lookup_df)} job codes")
    
    # Load template
    template_path = Path(r"C:\Users\krush\OneDrive - Walmart Inc\Documents\VSCode\Activity_Hub\Store Support\Projects\JobCodes-teaming\Job Codes\AMP Roles.xlsx.backup_before_update")
    
    if template_path.exists():
        template_df = pd.read_excel(template_path)
        print(f"✓ Loaded template with {len(template_df)} rows")
        
        # Merge with lookup
        if 'Current Job Code' in template_df.columns:
            job_code_col = 'Current Job Code'
        else:
            job_code_col = template_df.columns[0]
        
        # Rename for merge
        lookup_df_merge = lookup_df.copy()
        lookup_df_merge.rename(columns={'SMART Job Code': job_code_col}, inplace=True)
        
        # Merge
        result_df = template_df.merge(
            lookup_df_merge[[job_code_col, 'Role Type', 'User ID']],
            on=job_code_col,
            how='left'
        )
        
        print(f"✓ Merged data: {(result_df['User ID'].notna()).sum()} rows with User ID")
    else:
        print("⚠ Template not found, using lookup data directly")
        result_df = lookup_df
    
    # Save
    output_path = Path(r"C:\Users\krush\OneDrive - Walmart Inc\Documents\VSCode\Activity_Hub\Store Support\Projects\JobCodes-teaming\Job Codes\AMP Roles Updated.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    result_df.to_excel(output_path, sheet_name='AMP Roles', index=False)
    print(f"✓ File saved: {output_path}")
    
    file_size = output_path.stat().st_size
    print(f"✓ File size: {file_size:,} bytes")
    
    return True

if __name__ == "__main__":
    print("\n" + "="*70)
    print("AMP ROLES UPDATED - FRESH FILE CREATION")
    print("="*70)
    
    success = False
    
    # Try openpyxl first
    if use_openpyxl:
        try:
            success = create_with_openpyxl()
        except Exception as e:
            print(f"\n⚠ openpyxl method failed: {e}")
            print("Falling back to pandas...")
            success = create_with_pandas()
    else:
        success = create_with_pandas()
    
    if success:
        print("\n" + "="*70)
        print("✅ AMP ROLES UPDATED.XLSX CREATED SUCCESSFULLY")
        print("="*70)
        print("\nFile is ready to open and use!")
    else:
        print("\n❌ File creation failed")
        sys.exit(1)
