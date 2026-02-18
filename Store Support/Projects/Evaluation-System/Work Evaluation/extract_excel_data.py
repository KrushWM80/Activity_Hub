#!/usr/bin/env python3
import openpyxl
from collections import defaultdict
import json

excel_path = r"C:\Users\krush\Documents\VSCode\Kendalls Task and Work.xlsx"

try:
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Get headers
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    
    print("="*100)
    print("EXCEL FILE STRUCTURE")
    print("="*100)
    print(f"\nSheet Name: {ws.title}")
    print(f"Total Columns: {len(headers)}")
    print("\nHeaders:")
    for i, h in enumerate(headers):
        print(f"  [{i}] {h}")
    
    print("\n" + "="*100)
    print("DATA ROWS (First 100)")
    print("="*100)
    
    rows_data = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if any(cell is not None for cell in row):  # Skip completely empty rows
            rows_data.append(row)
            if len(rows_data) <= 100:
                print(f"\nRow {row_idx}:")
                for i, (header, value) in enumerate(zip(headers, row)):
                    if value is not None:
                        print(f"  {header}: {value}")
    
    print("\n" + "="*100)
    print(f"TOTAL DATA ROWS: {len(rows_data)}")
    print("="*100)
    
    # Try to categorize by project if there's a project column
    project_col = next((i for i, h in enumerate(headers) if h and 'project' in str(h).lower()), None)
    category_col = next((i for i, h in enumerate(headers) if h and 'category' in str(h).lower()), None)
    task_col = next((i for i, h in enumerate(headers) if h and 'task' in str(h).lower()), None)
    hours_col = next((i for i, h in enumerate(headers) if h and 'hours' in str(h).lower()), None)
    
    if project_col is not None:
        print("\nPROJECTS IDENTIFIED:")
        projects = defaultdict(list)
        for row in rows_data:
            if len(row) > project_col and row[project_col]:
                projects[row[project_col]].append(row)
        
        for proj, items in sorted(projects.items()):
            print(f"\n  {proj}: {len(items)} items")
    
    if hours_col is not None:
        print("\nHOURS IDENTIFIED:")
        total_hours = 0
        for row in rows_data:
            if len(row) > hours_col and row[hours_col]:
                try:
                    total_hours += float(row[hours_col])
                except:
                    pass
        print(f"  Total hours: {total_hours}")

except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
